from flask import Flask, render_template, flash, redirect, url_for, request
import os
import cv2
from openpyxl import load_workbook
import openpyxl
from datetime import date
import shutil

app = Flask(__name__)

app.config['SECRET_KEY'] = 'mysecretkey'

@app.route('/')
def index():
        with open('attendance/students.txt', 'r') as k:
                students = k.read().split()
        workbook = load_workbook(filename='attendance/attendance_register.xlsx')
        sheet = workbook.active
        return render_template('index.html', students=students, sheet=sheet)

@app.route('/collect_samples', methods = ['GET', 'POST'])
def collect_samples():
        if request.method == 'POST':
                username = request.form['Name']
                usn = request.form['USN']
                department = request.form['Department']
                sem = request.form['Sem']
        student_data = username + "." + usn + "." + department + "." + sem
        cmd = 'python attendance/image_collector.py ' + username
        with open('attendance/students.txt', 'r') as f:
                output = [line for line in f]
        for line in output:
                if line.split(".")[0] == username:
                        flash('already registered')
                        return redirect(url_for('index'))
        os.system(cmd)
        with open('attendance/students.txt', 'a') as j:
                j.write(student_data+'\n')
        workbook = load_workbook(filename='attendance/attendance_register.xlsx')
        sheet = workbook.active
        rc = 'A'+str(sheet.max_row+1)
        sheet[rc] = username
        workbook.save('attendance/attendance_register.xlsx')
        flash('samples collected')
        return redirect(url_for('index'))

@app.route('/recognize_face')
def recognize():
        with open('attendance/students.txt', 'r') as f:
                student = f.read()
        if not(student):
                flash('no student')
                return redirect(url_for('index'))
        cmd = 'python attendance/recognize_face.py'
        flash('recognized')
        os.system(cmd)
        student = None
        try:
                with open('attendance/username.txt', 'r') as f:
                        student = f.read()
        except:
                pass
        add_column()
        mark_present(student)
        return redirect(url_for('index'))

def add_column():
        workbook = load_workbook(filename='attendance/attendance_register.xlsx')
        sheet = workbook.active
        rc = str(sheet.cell(row=1,column=sheet.max_column))[-3:-1]
        print(rc,' rc',  sheet[rc].value, ' sheet[rc]', int(str(date.today()).split('-')[-1]), ' date.today')
        if sheet[rc].value != int(str(date.today()).split('-')[-1]):
                rc = str(sheet.cell(row=1,column=sheet.max_column+1))[-3:-1]
                sheet[rc] = int(str(date.today()).split('-')[-1])
        workbook.save('attendance/attendance_register.xlsx')

def mark_present(student):
        workbook = load_workbook(filename='attendance/attendance_register.xlsx')
        sheet = workbook.active
        column_no = str(sheet.cell(row=1,column=sheet.max_column))[-3:-2]
        for cell_no in sheet.iter_cols(min_row=2,max_row=sheet.max_row,min_col=1,max_col=1):
                names_row_cell_tuple=cell_no
        for row_cell_no in names_row_cell_tuple:
                if row_cell_no.value == student:
                        row_no = str(row_cell_no)[-2:-1]
        rc = column_no + row_no
        if sheet[rc].value != 'present':
                sheet[rc] = 'present'
        workbook.save('attendance/attendance_register.xlsx')

@app.route('/<username>')
def delete_student(username):
        with open('attendance/students.txt', 'r') as f:
                output = [line for line in f if not line.startswith(username)]
        with open('attendance/students.txt', 'w') as f:
                f.writelines(output)
        basedir = os.path.abspath(os.path.dirname(__file__))
        if os.path.exists(os.path.join(basedir, "static\\", username, "")):
                shutil.rmtree(os.path.join(basedir, "static\\", username, ""))
        workbook = load_workbook(filename='attendance/attendance_register.xlsx')
        sheet = workbook.active
        data = []
        for row in sheet.iter_rows(min_row=1, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
                sub_data = []
                if row[0].value != username:
                        for cell in row:
                                sub_data.append(cell.value)
                        data.append(tuple(sub_data))
        data = tuple(data)
        workbook.save('attendance/attendance_register.xlsx')
        os.remove('attendance/attendance_register.xlsx')
        wb = openpyxl.Workbook()
        sheet2 = wb.active
        for i in data:
                sheet2.append(i)
        wb.save('attendance/attendance_register.xlsx')
        return redirect(url_for('index'))